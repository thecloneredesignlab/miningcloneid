#!/usr/bin/env python3
"""Build joint soft-coupling warm-start tables from separate in vivo/in vitro fits.

Default inputs match the local 200-seed separate-fit result folders:

  oxygen/results/fit_invivo_O2G_buffering_200seed
  oxygen/results/fit_invitro_O2G_buffering_200seed

The script writes the four warm-start versions used for the 200-seed analysis:

  1. best in vivo seed, paired with the same seed's in vitro result
  2. best in vitro seed, paired with the same seed's in vivo result
  3. best in vivo seed plus best in vitro seed, using each context's own best seed
  4. paired-seed mean, averaging per-seed center/delta values on optimizer scale
"""

from __future__ import annotations

import argparse
import csv
import math
import re
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - runtime dependency check
    Workbook = None
    Font = None
    PatternFill = None
    get_column_letter = None


SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parents[3]

DEFAULT_INVIVO_DIR = PROJECT_ROOT / "oxygen/results/fit_invivo_O2G_buffering_200seed"
DEFAULT_INVITRO_DIR = PROJECT_ROOT / "oxygen/results/fit_invitro_O2G_buffering_200seed"
DEFAULT_OUT_DIR = PROJECT_ROOT / "oxygen/results/warm_start_from_separate_200seed"

TRANSFORMED_NAMES = {
    "lam_max": "log10_lam_max",
    "p_mis_base": "log10_p_mis_base",
    "p_misseg": "log10_p_misseg",
    "k_o_mis": "log10_k_o_mis",
    "buffer_smax": "buffer_smax",
    "buffer_beta": "log10_buffer_beta",
    "buffer_n_exp": "log10_buffer_n_exp",
    "p_wgd": "log10_p_wgd",
    "o2_S0": "log10_o2_S0",
    "kappa_O": "log10_kappa_O",
    "eta_o2": "log10_eta_o2",
    "rho_2N": "log10_rho_2N",
    "alpha_o2": "log10_alpha_o2",
    "gamma_growth": "gamma_growth",
    "mu_hp": "log10_mu_hp",
    "gamma_mu": "gamma_mu",
    "O2_crit": "log10_O2_crit",
    "n_O": "n_O",
    "k_clear": "log10_k_clear",
    "sigma_burden": "log10_sigma_burden",
}

TRANSFORMS = {
    "lam_max": "log10",
    "p_mis_base": "log10",
    "p_misseg": "log10",
    "k_o_mis": "log10",
    "buffer_smax": "identity",
    "buffer_beta": "log10",
    "buffer_n_exp": "log10",
    "p_wgd": "log10",
    "o2_S0": "log10",
    "kappa_O": "log10",
    "eta_o2": "log10",
    "rho_2N": "log10",
    "alpha_o2": "log10",
    "gamma_growth": "identity",
    "mu_hp": "log10",
    "gamma_mu": "identity",
    "O2_crit": "log10",
    "n_O": "identity",
    "k_clear": "log10",
    "sigma_burden": "log10",
}

WARM_START_ROWS = [
    ("mu_hp", "soft_coupled_center"),
    ("k_o_mis", "soft_coupled_center"),
    ("alpha_o2", "soft_coupled_center"),
    ("buffer_n_exp", "soft_coupled_center"),
    ("n_O", "soft_coupled_center"),
    ("O2_crit", "soft_coupled_center"),
    ("p_mis_base", "non_soft_shared_or_invivo"),
    ("lam_max", "non_soft_shared_or_invivo"),
    ("p_misseg", "soft_coupled_center"),
    ("buffer_beta", "soft_coupled_center"),
    ("buffer_smax", "soft_coupled_center"),
    ("gamma_mu", "non_soft_shared_or_invivo"),
    ("gamma_growth", "soft_coupled_center"),
    ("p_wgd", "non_soft_shared_or_invivo"),
]

EXTRA_INVIVO_INIT_PARAMS = [
    "o2_S0",
    "kappa_O",
    "eta_o2",
    "rho_2N",
    "k_clear",
    "sigma_burden",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--invivo-dir", type=Path, default=DEFAULT_INVIVO_DIR)
    parser.add_argument("--invitro-dir", type=Path, default=DEFAULT_INVITRO_DIR)
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    parser.add_argument(
        "--run-label",
        default="",
        help="Label used in combined workbook names, for example 200seed or 500seed. Default infers from input/output paths.",
    )
    parser.add_argument(
        "--seed-list",
        default="",
        help="Optional comma-separated seed list. Default uses seed dirs common to both result folders.",
    )
    parser.add_argument("--no-xlsx", action="store_true", help="Write TSV/CSV only.")
    return parser.parse_args()


def read_key_value_tsv(path: Path, key_col: str, value_col: str) -> dict[str, str]:
    with path.open(newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        return {
            str(row[key_col]).strip(): str(row[value_col]).strip()
            for row in reader
            if str(row.get(key_col, "")).strip()
        }


def as_float(x: object) -> float:
    try:
        return float(x)
    except (TypeError, ValueError):
        return math.nan


def read_params(seed_dir: Path) -> dict[str, float]:
    raw = read_key_value_tsv(seed_dir / "best_params.tsv", "parameter", "value")
    return {key: as_float(value) for key, value in raw.items()}


def read_summary(seed_dir: Path) -> dict[str, str]:
    return read_key_value_tsv(seed_dir / "fit_summary.tsv", "metric", "value")


def seed_dirs(parent: Path) -> dict[int, Path]:
    out: dict[int, Path] = {}
    pattern = re.compile(r"^seed(\d+)$")
    for child in parent.iterdir():
        if not child.is_dir():
            continue
        match = pattern.match(child.name)
        if not match:
            continue
        if (child / "best_params.tsv").exists() and (child / "fit_summary.tsv").exists():
            out[int(match.group(1))] = child
    return out


def objective_value(summary: dict[str, str], mode: str) -> float:
    keys = (
        ["objective", "optimizer_local_objective", "optimizer_deoptim_objective"]
        if mode == "invivo"
        else ["objective_total", "optimizer_local_objective", "optimizer_deoptim_objective"]
    )
    for key in keys:
        value = as_float(summary.get(key))
        if math.isfinite(value):
            return value
    return math.nan


def transform_value(value: float, transform: str) -> float:
    if transform == "identity":
        return value
    if transform == "log10":
        return math.log10(value) if math.isfinite(value) and value > 0 else math.nan
    raise ValueError(f"Unsupported transform: {transform}")


def inverse_transform(value: float, transform: str) -> float:
    if not math.isfinite(value):
        return math.nan
    if transform == "identity":
        return value
    if transform == "log10":
        return 10**value
    raise ValueError(f"Unsupported transform: {transform}")


def numeric_mean(values: list[float]) -> float:
    ok = [value for value in values if math.isfinite(value)]
    return sum(ok) / len(ok) if ok else math.nan


def single_pair_transformed_rows(
    invivo_seed: int,
    invitro_seed: int,
    invivo_params: dict[str, float],
    invitro_params: dict[str, float],
) -> list[dict[str, float | str | int]]:
    rows = []
    for parameter, role in WARM_START_ROWS:
        transform = TRANSFORMS[parameter]
        vivo_t = transform_value(invivo_params.get(parameter, math.nan), transform)
        vitro_t = transform_value(invitro_params.get(parameter, math.nan), transform)
        rows.append(
            {
                "parameter": parameter,
                "role": role,
                "transform": transform,
                "joint_init_name": TRANSFORMED_NAMES[parameter],
                "vivo_transformed": vivo_t,
                "vitro_transformed": vitro_t,
                "center_transformed": (vivo_t + vitro_t) / 2.0,
                "delta_transformed": vivo_t - vitro_t,
                "source_invivo_seed": invivo_seed,
                "source_invitro_seed": invitro_seed,
            }
        )
    return rows


def warm_rows_from_transformed(
    transformed_rows: list[dict[str, float | str | int]],
    *,
    version: str,
    source_seed: int | str,
    n_seed_pairs: int,
    aggregation: str,
    include_context_seed_cols: bool = False,
) -> list[dict[str, float | str | int]]:
    out = []
    for base in transformed_rows:
        parameter = str(base["parameter"])
        transform = str(base["transform"])
        center_t = float(base["center_transformed"])
        delta_t = float(base["delta_transformed"])
        vivo_t = center_t + delta_t / 2.0
        vitro_t = center_t - delta_t / 2.0
        vivo_nat = inverse_transform(vivo_t, transform)
        vitro_nat = inverse_transform(vitro_t, transform)
        ratio = vivo_nat / vitro_nat if math.isfinite(vitro_nat) and vitro_nat != 0 else math.nan
        row = {
            "parameter": parameter,
            "role": str(base["role"]),
            "transform": transform,
            "joint_init_name": str(base["joint_init_name"]),
            "vivo_natural": vivo_nat,
            "vitro_natural": vitro_nat,
            "ratio_vivo_to_vitro": ratio,
            "center_transformed": center_t,
            "delta_transformed": delta_t,
            "center_natural": inverse_transform(center_t, transform),
            "joint_init_value": center_t,
            "rank": abs(1.0 - ratio) if math.isfinite(ratio) else math.nan,
            "source_seed": source_seed,
        }
        if include_context_seed_cols:
            row["source_invivo_seed"] = int(base["source_invivo_seed"])
            row["source_invitro_seed"] = int(base["source_invitro_seed"])
        row.update(
            {
                "n_seed_pairs": n_seed_pairs,
                "aggregation": aggregation,
                "version": version,
            }
        )
        out.append(row)
    return out


def build_single_seed_warm(
    seed: int,
    invivo_params: dict[str, float],
    invitro_params: dict[str, float],
    *,
    version: str,
) -> list[dict[str, float | str | int]]:
    return warm_rows_from_transformed(
        single_pair_transformed_rows(seed, seed, invivo_params, invitro_params),
        version=version,
        source_seed=seed,
        n_seed_pairs=1,
        aggregation="single_paired_seed",
    )


def build_best_each_warm(
    invivo_seed: int,
    invitro_seed: int,
    invivo_params: dict[str, float],
    invitro_params: dict[str, float],
) -> list[dict[str, float | str | int]]:
    return warm_rows_from_transformed(
        single_pair_transformed_rows(invivo_seed, invitro_seed, invivo_params, invitro_params),
        version="best_each_seed",
        source_seed=f"invivo{invivo_seed}_invitro{invitro_seed}",
        n_seed_pairs=2,
        aggregation="best_invivo_seed_plus_best_invitro_seed",
        include_context_seed_cols=True,
    )


def build_paired_mean_warm(
    common_seeds: list[int],
    invivo_by_seed: dict[int, dict[str, float]],
    invitro_by_seed: dict[int, dict[str, float]],
) -> list[dict[str, float | str | int]]:
    per_seed = []
    for seed in common_seeds:
        per_seed.extend(single_pair_transformed_rows(seed, seed, invivo_by_seed[seed], invitro_by_seed[seed]))
    mean_rows = []
    for parameter, role in WARM_START_ROWS:
        subset = [row for row in per_seed if row["parameter"] == parameter]
        mean_rows.append(
            {
                "parameter": parameter,
                "role": role,
                "transform": TRANSFORMS[parameter],
                "joint_init_name": TRANSFORMED_NAMES[parameter],
                "center_transformed": numeric_mean([float(row["center_transformed"]) for row in subset]),
                "delta_transformed": numeric_mean([float(row["delta_transformed"]) for row in subset]),
            }
        )
    return warm_rows_from_transformed(
        mean_rows,
        version="paired_seed_mean",
        source_seed="",
        n_seed_pairs=len(common_seeds),
        aggregation="mean_of_per_seed_transformed_pairs",
    )


def invivo_extra_init_rows(
    invivo_params_list: list[dict[str, float]],
    *,
    version: str,
    source_seed: int | str,
    n_seed_pairs: int,
    aggregation: str,
    context_seed_cols: dict[str, int | str] | None = None,
) -> list[dict[str, float | str | int]]:
    rows = []
    context_seed_cols = context_seed_cols or {}
    for parameter in EXTRA_INVIVO_INIT_PARAMS:
        if parameter not in TRANSFORMED_NAMES:
            continue
        vals_t = [
            transform_value(params.get(parameter, math.nan), TRANSFORMS[parameter])
            for params in invivo_params_list
        ]
        rows.append(
            {
                "joint_init_name": TRANSFORMED_NAMES[parameter],
                "init_value": numeric_mean(vals_t),
                "source_role": "non_soft_invivo_only",
                "parameter": parameter,
                "transform": TRANSFORMS[parameter],
                "source_seed": source_seed,
                "n_seed_pairs": n_seed_pairs,
                "aggregation": aggregation,
                "version": version,
                **context_seed_cols,
            }
        )
    return rows


def build_joint_init_values(
    warm_rows: list[dict[str, float | str | int]],
    invivo_params_list: list[dict[str, float]],
    *,
    version: str,
    source_seed: int | str,
    n_seed_pairs: int,
    aggregation: str,
) -> list[dict[str, float | str | int]]:
    out = []
    context_seed_cols: dict[str, int | str] = {}
    if warm_rows and "source_invivo_seed" in warm_rows[0]:
        context_seed_cols = {
            "source_invivo_seed": warm_rows[0].get("source_invivo_seed", ""),
            "source_invitro_seed": warm_rows[0].get("source_invitro_seed", ""),
        }
    for row in warm_rows:
        out.append(
            {
                "joint_init_name": row["joint_init_name"],
                "init_value": row["joint_init_value"],
                "source_role": row["role"],
                "parameter": row["parameter"],
                "transform": row["transform"],
                "source_seed": source_seed,
                "n_seed_pairs": n_seed_pairs,
                "aggregation": aggregation,
                "version": version,
                **context_seed_cols,
            }
        )
        if row["role"] == "soft_coupled_center":
            out.append(
                {
                    "joint_init_name": f"delta__{row['joint_init_name']}",
                    "init_value": row["delta_transformed"],
                    "source_role": "soft_coupled_delta",
                    "parameter": row["parameter"],
                    "transform": row["transform"],
                    "source_seed": source_seed,
                    "n_seed_pairs": n_seed_pairs,
                    "aggregation": aggregation,
                    "version": version,
                    **context_seed_cols,
                }
            )
    out.extend(
        invivo_extra_init_rows(
            invivo_params_list,
            version=version,
            source_seed=source_seed,
            n_seed_pairs=n_seed_pairs,
            aggregation=aggregation,
            context_seed_cols=context_seed_cols,
        )
    )
    return out


def write_tsv(path: Path, rows: list[dict[str, object]]) -> None:
    if not rows:
        path.write_text("")
        return
    cols = list(rows[0].keys())
    with path.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=cols, delimiter="\t", lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def add_sheet(workbook, name: str, rows: list[dict[str, object]]) -> None:
    sheet = workbook.create_sheet(title=name[:31])
    if not rows:
        return
    cols = list(rows[0].keys())
    sheet.append(cols)
    for row in rows:
        sheet.append([row.get(col, "") for col in cols])
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    sheet.freeze_panes = "A2"
    for idx, col in enumerate(cols, start=1):
        max_len = max(len(str(col)), *(len(str(cell.value)) for cell in sheet[get_column_letter(idx)] if cell.value is not None))
        sheet.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 34)


def write_workbook(path: Path, sheets: dict[str, list[dict[str, object]]]) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl is required for xlsx output. Re-run with --no-xlsx to write TSV only.")
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, rows in sheets.items():
        add_sheet(workbook, name, rows)
    workbook.save(path)


def parse_seed_list(seed_list: str, invivo_dirs: dict[int, Path], invitro_dirs: dict[int, Path]) -> list[int]:
    if seed_list.strip():
        seeds = [int(x) for x in re.split(r"[,; ]+", seed_list.strip()) if x]
        missing = [seed for seed in seeds if seed not in invivo_dirs or seed not in invitro_dirs]
        if missing:
            raise SystemExit(f"Seeds missing from one or both result directories: {missing}")
        return sorted(set(seeds))
    return sorted(set(invivo_dirs).intersection(invitro_dirs))


def infer_run_label(invivo_dir: Path, invitro_dir: Path, out_dir: Path, n_seeds: int) -> str:
    text = " ".join(str(path) for path in (out_dir, invivo_dir, invitro_dir))
    matches = re.findall(r"([0-9]+seed)", text)
    if matches:
        return matches[0]
    return f"{n_seeds}seed"


def main() -> None:
    args = parse_args()
    invivo_dir = args.invivo_dir.resolve()
    invitro_dir = args.invitro_dir.resolve()
    out_dir = args.out_dir.resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    invivo_dirs = seed_dirs(invivo_dir)
    invitro_dirs = seed_dirs(invitro_dir)
    common_seeds = parse_seed_list(args.seed_list, invivo_dirs, invitro_dirs)
    if not common_seeds:
        raise SystemExit("No paired seeds found.")
    run_label = args.run_label.strip() or infer_run_label(invivo_dir, invitro_dir, out_dir, len(common_seeds))

    invivo_summary = {seed: read_summary(invivo_dirs[seed]) for seed in common_seeds}
    invitro_summary = {seed: read_summary(invitro_dirs[seed]) for seed in common_seeds}
    invivo_params = {seed: read_params(invivo_dirs[seed]) for seed in common_seeds}
    invitro_params = {seed: read_params(invitro_dirs[seed]) for seed in common_seeds}

    seed_summary = []
    for seed in common_seeds:
        seed_summary.append(
            {
                "seed": seed,
                "invivo_objective": objective_value(invivo_summary[seed], "invivo"),
                "invitro_objective_total": objective_value(invitro_summary[seed], "invitro"),
                "invivo_optimizer_local_objective": as_float(invivo_summary[seed].get("optimizer_local_objective")),
                "invitro_optimizer_local_objective": as_float(invitro_summary[seed].get("optimizer_local_objective")),
                "invivo_seed_dir": str(invivo_dirs[seed]),
                "invitro_seed_dir": str(invitro_dirs[seed]),
            }
        )

    best_invivo_seed = min(
        common_seeds,
        key=lambda seed: (
            objective_value(invivo_summary[seed], "invivo")
            if math.isfinite(objective_value(invivo_summary[seed], "invivo"))
            else math.inf,
            seed,
        ),
    )
    best_invitro_seed = min(
        common_seeds,
        key=lambda seed: (
            objective_value(invitro_summary[seed], "invitro")
            if math.isfinite(objective_value(invitro_summary[seed], "invitro"))
            else math.inf,
            seed,
        ),
    )

    best_invivo_warm = build_single_seed_warm(
        best_invivo_seed,
        invivo_params[best_invivo_seed],
        invitro_params[best_invivo_seed],
        version="best_invivo_seed",
    )
    best_invivo_init = build_joint_init_values(
        best_invivo_warm,
        [invivo_params[best_invivo_seed]],
        version="best_invivo_seed",
        source_seed=best_invivo_seed,
        n_seed_pairs=1,
        aggregation="single_paired_seed",
    )

    best_invitro_warm = build_single_seed_warm(
        best_invitro_seed,
        invivo_params[best_invitro_seed],
        invitro_params[best_invitro_seed],
        version="best_invitro_seed",
    )
    best_invitro_init = build_joint_init_values(
        best_invitro_warm,
        [invivo_params[best_invitro_seed]],
        version="best_invitro_seed",
        source_seed=best_invitro_seed,
        n_seed_pairs=1,
        aggregation="single_paired_seed",
    )

    best_each_warm = build_best_each_warm(
        best_invivo_seed,
        best_invitro_seed,
        invivo_params[best_invivo_seed],
        invitro_params[best_invitro_seed],
    )
    best_each_init = build_joint_init_values(
        best_each_warm,
        [invivo_params[best_invivo_seed]],
        version="best_each_seed",
        source_seed=f"invivo{best_invivo_seed}_invitro{best_invitro_seed}",
        n_seed_pairs=2,
        aggregation="best_invivo_seed_plus_best_invitro_seed",
    )

    paired_mean_warm = build_paired_mean_warm(common_seeds, invivo_params, invitro_params)
    paired_mean_init = build_joint_init_values(
        paired_mean_warm,
        [invivo_params[seed] for seed in common_seeds],
        version="paired_seed_mean",
        source_seed="",
        n_seed_pairs=len(common_seeds),
        aggregation="mean_of_per_seed_transformed_pairs",
    )

    metadata = [
        {"key": "invivo_dir", "value": str(invivo_dir)},
        {"key": "invitro_dir", "value": str(invitro_dir)},
        {"key": "n_paired_seeds", "value": len(common_seeds)},
        {"key": "paired_seeds", "value": ",".join(str(seed) for seed in common_seeds)},
        {"key": "best_invivo_seed", "value": best_invivo_seed},
        {"key": "best_invivo_objective_metric", "value": "objective"},
        {"key": "best_invivo_objective", "value": objective_value(invivo_summary[best_invivo_seed], "invivo")},
        {"key": "best_invitro_seed", "value": best_invitro_seed},
        {"key": "best_invitro_objective_metric", "value": "objective_total"},
        {"key": "best_invitro_objective_total", "value": objective_value(invitro_summary[best_invitro_seed], "invitro")},
        {"key": "best_each_version", "value": f"best_each_seed{best_invivo_seed}_invivo_seed{best_invitro_seed}_invitro"},
        {"key": "best_each_invivo_seed", "value": best_invivo_seed},
        {"key": "best_each_invitro_seed", "value": best_invitro_seed},
        {"key": "best_each_rule", "value": "vivo values from best in vivo seed; vitro values from best in vitro seed"},
        {"key": "paired_mean_rule", "value": "mean center/delta/init values on transformed optimizer scale, then inverse-transform natural columns"},
        {"key": "soft_coupled_rows", "value": ",".join(p for p, role in WARM_START_ROWS if role == "soft_coupled_center")},
        {"key": "non_soft_rows", "value": ",".join(p for p, role in WARM_START_ROWS if role != "soft_coupled_center")},
    ]

    write_tsv(out_dir / "metadata.tsv", metadata)
    write_tsv(out_dir / "seed_selection_summary.tsv", seed_summary)

    variants = [
        (f"best_invivo_seed{best_invivo_seed}", best_invivo_warm, best_invivo_init),
        (f"best_invitro_seed{best_invitro_seed}", best_invitro_warm, best_invitro_init),
        (f"best_each_seed{best_invivo_seed}_invivo_seed{best_invitro_seed}_invitro", best_each_warm, best_each_init),
        (f"paired_mean_{len(common_seeds)}seed", paired_mean_warm, paired_mean_init),
    ]
    for label, warm_rows, init_rows in variants:
        write_tsv(out_dir / f"{label}_warm_start_table.tsv", warm_rows)
        write_tsv(out_dir / f"{label}_joint_init_values.tsv", init_rows)
        if not args.no_xlsx:
            write_workbook(
                out_dir / f"{label}.xlsx",
                {
                    "warm_start_table": warm_rows,
                    "joint_init_values": init_rows,
                    "metadata": metadata,
                    "seed_selection": seed_summary,
                },
            )

    if not args.no_xlsx:
        write_workbook(
            out_dir / f"joint_soft_coupling_warm_start_{run_label}_three_versions.xlsx",
            {
                "metadata": metadata,
                "seed_selection": seed_summary,
                f"best_invivo_s{best_invivo_seed}_warm": best_invivo_warm,
                f"best_invivo_s{best_invivo_seed}_init": best_invivo_init,
                f"best_invitro_s{best_invitro_seed}_warm": best_invitro_warm,
                f"best_invitro_s{best_invitro_seed}_init": best_invitro_init,
                f"paired_mean_{len(common_seeds)}s_warm": paired_mean_warm,
                f"paired_mean_{len(common_seeds)}s_init": paired_mean_init,
            },
        )
        write_workbook(
            out_dir / f"joint_soft_coupling_warm_start_{run_label}_four_versions.xlsx",
            {
                "metadata": metadata,
                "seed_selection": seed_summary,
                f"best_invivo_s{best_invivo_seed}_warm": best_invivo_warm,
                f"best_invivo_s{best_invivo_seed}_init": best_invivo_init,
                f"best_invitro_s{best_invitro_seed}_warm": best_invitro_warm,
                f"best_invitro_s{best_invitro_seed}_init": best_invitro_init,
                f"best_each_s{best_invivo_seed}_{best_invitro_seed}_warm": best_each_warm,
                f"best_each_s{best_invivo_seed}_{best_invitro_seed}_init": best_each_init,
                f"paired_mean_{len(common_seeds)}s_warm": paired_mean_warm,
                f"paired_mean_{len(common_seeds)}s_init": paired_mean_init,
            },
        )

    print(f"paired_seeds={len(common_seeds)}")
    print(f"best_invivo_seed={best_invivo_seed} objective={objective_value(invivo_summary[best_invivo_seed], 'invivo')}")
    print(f"best_invitro_seed={best_invitro_seed} objective_total={objective_value(invitro_summary[best_invitro_seed], 'invitro')}")
    print(f"out_dir={out_dir}")


if __name__ == "__main__":
    main()
